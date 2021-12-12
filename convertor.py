import csv
from datetime import datetime

from django.core.files import File

from openpyxl import load_workbook


class Gsuite:
    def __init__(self):
        self.students = []
        self.s_code = ""
        self.yearCode = ""
        self.grade = ""
        self.class_ = ""
        self.s_num = 0
        self.samplePW = ""
        self.orgPath = ""
        self.s_admin = ""

    def init_data(self):
        # data for school_user.csv
        self.data = []
        with open("staticfiles/G-suite/users.csv", newline="") as f:
            r = csv.reader(f, delimiter=",")
            for row in r:
                self.data.append(row)

    def put_data(self):
        temp = ["" for x in range(25)] + ["TRUE", ""]
        for i, name in enumerate(self.students):
            s_num = i + 1
            print(name)
            firstName = name[1:]
            lastName = name[0]

            email = f"{self.s_code}s-{self.yearCode}{self.grade}{self.class_.zfill(2)}{s_num:02d}@cberi.go.kr"

            # put data in temp
            temp[0] = firstName
            temp[1] = lastName
            temp[2] = email
            temp[3] = self.samplePW
            temp[5] = self.orgPath
            self.data.append(temp.copy())

    def make_csv(self):
        # save result csv file
        self.file_name = f"{self.s_admin}{self.grade}{self.class_.zfill(2)}_user.csv"
        self.result_file = f"staticfiles/G-suite/results/{self.file_name}"
        with open(self.result_file, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)
            for row in self.data:
                writer.writerow(row)

    def get_url(self):
        with open(self.result_file, "r") as f:
            tmpfile = File(f)
            self.file_url = "/" + tmpfile.name


class GUser(Gsuite):
    def __init__(self, file, s_info, grade, classN):
        self.get_properties(file, s_info, grade, classN)
        self.init_data()
        self.put_data()
        self.make_csv()
        self.get_url()

    def get_properties(self, file, s_info, grade, classN):
        # xlsx 파일에서 학생정보 가져오기
        wb = load_workbook(file)
        ws = wb[wb.sheetnames[0]]
        for max_row, row in enumerate(ws, 1):
            if all(c.value is None for c in row):
                break
        print("ws.max row:", ws.max_row)
        max_row -= 1
        print("max row:", max_row)
        self.students = []
        for r in range(2, max_row + 1):
            self.students.append(ws.cell(row=r, column=2).value)

        print(self.students)
        # Properties
        self.s_code, self.school, self.s_admin = s_info
        self.grade = grade
        self.class_ = classN

        thisYear = datetime.now().strftime("%Y")
        self.yearCode = thisYear[2:]

        self.samplePW = f"{self.s_admin}{self.s_code}"
        if len(self.samplePW) < 8:
            self.samplePW += "!"
        self.orgPath = f"/{self.school}"


class GUser_school(Gsuite):
    def __init__(self, file, s_info):
        self.get_properties(file, s_info)
        self.init_data()
        self.put_data()
        self.make_csv()
        self.get_url()

    def get_properties(self, file, s_info):
        # Properties
        self.s_code, self.school, self.s_admin = s_info

        thisYear = datetime.now().strftime("%Y")
        self.yearCode = thisYear[2:]

        self.samplePW = f"{self.s_admin}{self.s_code}"
        while len(self.samplePW) < 8:
            self.samplePW += "!"
        self.orgPath = f"/{self.school}"

        # xlsx 파일에서 학생정보 가져오기
        # 명렬표 불러오기
        wb = load_workbook(file)
        ws = wb.active

        for max_row, row in enumerate(ws, 1):
            if all(c.value is None for c in row):
                break
        print("ws.max row:", ws.max_row)
        max_row -= 1
        print("max row:", max_row)

        # 학년, 반, 이름 정보가져오기
        self.grade = []
        self.class_ = []
        self.s_num = []
        self.students = []
        for r in range(2, max_row + 1):
            self.grade.append(ws.cell(row=r, column=1).value)
            self.class_.append(ws.cell(row=r, column=2).value)
            self.s_num.append(ws.cell(row=r, column=3).value)
            self.students.append(ws.cell(row=r, column=4).value)

    def put_data(self):
        temp = ["" for x in range(25)] + ["TRUE", ""]
        for i, name in enumerate(self.students):
            grade = self.grade[i]
            class_ = self.class_[i]
            s_num = self.s_num[i]

            firstName = name[1:]
            lastName = name[0]

            email = f"{self.s_code}s-{self.yearCode}{grade}{class_.zfill(2)}{s_num.zfill(2)}@cberi.go.kr"

            # put data in temp
            temp[0] = firstName
            temp[1] = lastName
            temp[2] = email
            temp[3] = self.samplePW
            temp[5] = self.orgPath
            self.data.append(temp.copy())

    def make_csv(self):
        # save result csv file
        self.file_name = f"{self.s_admin}_user.csv"
        self.result_file = f"staticfiles/G-suite/results/{self.file_name}"
        with open(self.result_file, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)
            for row in self.data:
                writer.writerow(row)
